import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import datetime

# Function to load data from Excel
def load_data(file_path):
    df = pd.read_excel(file_path)
    return df

# Function to find hospitals with the required blood units
def find_hospitals(df, blood_type, required_units):
    available_hospitals = df[df[blood_type] >= required_units]
    return available_hospitals[['Hospital Name', blood_type]]

# Function to insert a booking into Excel
def insert_booking(file_path, hospital_name, blood_type, booked_units):
    wb = load_workbook(file_path)
    ws = wb.active
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=hospital_name)
    ws.cell(row=next_row, column=2, value=blood_type)
    ws.cell(row=next_row, column=3, value=booked_units)
    ws.cell(row=next_row, column=4, value=datetime.datetime.now())
    wb.save(file_path)

# Function to update blood units in Excel
def update_blood_units(file_path, hospital_name, blood_type, booked_units):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == hospital_name:
            blood_type_col = [cell.column for cell in ws[1] if cell.value == blood_type][0]
            total_units_col = [cell.column for cell in ws[1] if cell.value == 'Total Units'][0]
            blood_type_value = row[blood_type_col - 1].value
            total_units_value = row[total_units_col - 1].value
            
            if blood_type_value is not None and total_units_value is not None:
                row[blood_type_col - 1].value = max(0, blood_type_value - booked_units)
                row[total_units_col - 1].value = max(0, total_units_value - booked_units)
                row[-1].value = datetime.datetime.now()
            break
    wb.save(file_path)

# Path to your Excel file
file_path = os.path.abspath('blood_inventory.xlsx')

# Load the data
df = load_data(file_path)

# Streamlit app
st.title('Blood Unit Availability and Booking System')

# Initialize session state
if 'available_hospitals' not in st.session_state:
    st.session_state.available_hospitals = None
if 'selected_hospital' not in st.session_state:
    st.session_state.selected_hospital = None
if 'blood_type' not in st.session_state:
    st.session_state.blood_type = None
if 'required_units' not in st.session_state:
    st.session_state.required_units = 0

# Select blood type
st.session_state.blood_type = st.selectbox('Select Blood Type', df.columns[1:9].tolist(), index=df.columns[1:9].tolist().index(st.session_state.blood_type) if st.session_state.blood_type else 0)

# Input required units
st.session_state.required_units = st.number_input('Required Units', min_value=0, value=st.session_state.required_units)

# Search button
if st.button('Search'):
    df = load_data(file_path)  # Load data each time search is clicked
    st.session_state.available_hospitals = find_hospitals(df, st.session_state.blood_type, st.session_state.required_units)
    
    if not st.session_state.available_hospitals.empty:
        st.write('Hospitals with the required blood units:')
        st.write(st.session_state.available_hospitals)
    else:
        st.write('No hospitals found with the required blood units.')

# Show the booking form if hospitals are available
if st.session_state.available_hospitals is not None and not st.session_state.available_hospitals.empty:
    st.subheader('Book Blood Units')
    st.session_state.selected_hospital = st.selectbox('Select Hospital', st.session_state.available_hospitals['Hospital Name'].tolist(), index=st.session_state.available_hospitals['Hospital Name'].tolist().index(st.session_state.selected_hospital) if st.session_state.selected_hospital else 0)
    
    max_units = int(df[df['Hospital Name'] == st.session_state.selected_hospital][st.session_state.blood_type].values[0])
    booked_units = st.number_input('Units to Book', min_value=1, max_value=max_units, key='book_units')

    if st.button('Book'):
        if booked_units <= max_units:
            insert_booking(file_path, st.session_state.selected_hospital, st.session_state.blood_type, booked_units)
            update_blood_units(file_path, st.session_state.selected_hospital, st.session_state.blood_type, booked_units)
            st.write('Booking successful!')
            df = load_data(file_path)  # Reload data after booking
            # Reset session state after booking
            st.session_state.available_hospitals = None
            st.session_state.selected_hospital = None
            st.session_state.required_units = 0
        else:
            st.write(f'Cannot book {booked_units} units. Maximum available units: {max_units}')
else:
    if st.session_state.available_hospitals is not None:
        st.write('No hospitals found with the required blood units.')

# Optionally, add functionality to update the data
if st.sidebar.button("Reload Data"):
    df = load_data(file_path)
    st.sidebar.write("Data reloaded!")
